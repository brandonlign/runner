#!/usr/bin/env python3
"""Documentation-only audit of the public UKMON data dictionary; no meteor API call."""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook

TERMS = {
    "solar_longitude": [r"solar", r"sol"],
    "radiant_ra": [r"right ascension", r"radiant.*ra", r"\bra\b"],
    "radiant_dec": [r"declination", r"radiant.*dec", r"\bdec\b"],
    "geocentric_speed": [r"geocentric.*(speed|velocity)", r"\bvg\b", r"v[_ ]?g"],
    "perihelion_q": [r"perihelion", r"\bq\b"],
    "eccentricity": [r"eccentricity", r"\be\b"],
    "inclination": [r"inclination", r"\bi\b"],
    "argument_perihelion": [r"argument.*peri", r"omega", r"arg"],
    "ascending_node": [r"ascending.*node", r"node", r"omega"],
    "trajectory_id": [r"trajectory", r"orbit.*name", r"identifier", r"orbname"],
}


def sha256(path: Path) -> str:
    h=hashlib.sha256()
    with path.open('rb') as fh:
        for chunk in iter(lambda:fh.read(1024*1024),b''):
            h.update(chunk)
    return h.hexdigest()


def main()->int:
    p=argparse.ArgumentParser()
    p.add_argument('--dictionary',required=True,type=Path)
    p.add_argument('--freshness-json',required=True,type=Path)
    p.add_argument('--output',required=True,type=Path)
    a=p.parse_args(); a.output.mkdir(parents=True,exist_ok=True)

    fresh=json.loads(a.freshness_json.read_text())
    if fresh['verdict']!='PASS_UKMON_2024_2025_REPO_SCIENTIFIC_FRESHNESS_AUDIT': raise RuntimeError('freshness prerequisite did not pass')
    if fresh['potential_exposure_hit_count']!=0: raise RuntimeError('freshness exposure hit appeared')

    wb=load_workbook(a.dictionary,read_only=True,data_only=True)
    rows=[]
    for ws in wb.worksheets:
        for row_idx,row in enumerate(ws.iter_rows(values_only=True),start=1):
            values=["" if value is None else str(value) for value in row]
            text=" | ".join(values)
            lowered=text.lower()
            matched=[]
            for concept,patterns in TERMS.items():
                if any(re.search(pattern,lowered,re.I) for pattern in patterns): matched.append(concept)
            if matched:
                rows.append({'sheet':ws.title,'row':row_idx,'values':values,'matched_concepts':matched})

    concepts={concept:[] for concept in TERMS}
    for row in rows:
        for concept in row['matched_concepts']:
            concepts[concept].append({'sheet':row['sheet'],'row':row['row'],'values':row['values']})

    result={
        'verdict':'PASS_UKMON_DOCUMENTATION_INTERFACE_AUDIT' if all(concepts[c] for c in ('solar_longitude','radiant_ra','radiant_dec','geocentric_speed','trajectory_id')) else 'FAIL_UKMON_DOCUMENTATION_INTERFACE_AUDIT',
        'dictionary_sha256':sha256(a.dictionary),
        'dictionary_bytes':a.dictionary.stat().st_size,
        'sheet_names':wb.sheetnames,
        'concept_matches':concepts,
        'meteor_api_contacted':False,
        'meteor_record_access':False,
        'reserved_2024_2025_access':False,
        'target_information_access':False,
        'claim_boundary':'Only the UKMON-published data-dictionary workbook was downloaded and inspected as documentation metadata. No matches/trajectory API endpoint or meteor record was requested.',
    }
    (a.output/'ukmon_documentation_interface_audit.json').write_text(json.dumps(result,indent=2,sort_keys=True)+'\n')
    print(json.dumps(result,indent=2,sort_keys=True))
    if result['verdict'].startswith('FAIL_'): raise SystemExit(1)
    return 0

if __name__=='__main__': raise SystemExit(main())
